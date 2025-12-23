### 2025/12/23 test ok
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import os
import time
import sys
import re
from dotenv import load_dotenv
from datetime import datetime
from openpyxl.styles import NamedStyle

load_dotenv()

scope = ["https://spreadsheets.google.com/feeds", 'https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name('TOUR_JSON_FILE_NAME.json', scope)
client = gspread.authorize(creds)
sheet = client.open("Bank").worksheet("總明細")

HEADLESS = False

chrome_options = Options()
if HEADLESS:
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--window-size=1920,1080")
else:
    chrome_options.add_argument("--start-maximized")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--disable-blink-features=AutomationControlled")
driver = webdriver.Chrome(options=chrome_options)

class Bank:
    def __init__(self):
        self.login_id = 0
        self.login_account = 0
        self.login_password = 0
        self.main_account = 0
        self.cash = 0
        self.exchange = 0
        self.stock = 0

Esun = Bank()
Esun.login_id = os.getenv("ESUN_ID")
Esun.login_account = os.getenv("ESUN_ACCOUNT")
Esun.login_password = os.getenv("ESUN_PASSWORD")

Cathay = Bank()
Cathay.login_id = os.getenv("CATHAY_ID")
Cathay.login_account = os.getenv("CATHAY_ACCOUNT")
Cathay.login_password = os.getenv("CATHAY_PASSWORD")

Line = Bank()
Line.login_id = os.getenv("LINE_ID")
Line.login_account = os.getenv("LINE_ACCOUNT")
Line.login_password = os.getenv("LINE_PASSWORD")


def EsunSpider():
    driver.get("https://ebank.esunbank.com.tw/index.jsp")
    wait = WebDriverWait(driver, 20)

    driver.switch_to.default_content()
    WebDriverWait(driver, 20).until(
        EC.frame_to_be_available_and_switch_to_it((By.ID, "iframe1"))
    )

    cust_input = WebDriverWait(driver, 20).until(
        EC.visibility_of_element_located((By.ID, "loginform:custid"))
    )
    cust_input.clear()
    cust_input.send_keys(Esun.login_id)

    cust_input = WebDriverWait(driver, 20).until(
        EC.visibility_of_element_located((By.ID, "loginform:name"))
    )
    cust_input.clear()
    cust_input.send_keys(Esun.login_account)

    cust_input = WebDriverWait(driver, 20).until(
        EC.visibility_of_element_located((By.ID, "loginform:pxsswd"))
    )
    cust_input.clear()
    cust_input.send_keys(Esun.login_password)

    login_btn = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.ID, "loginform:linkCommand"))
    )
    login_btn.click()

    span_el = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID, "_0"))
    )
    Esun.main_account = span_el.text.strip()
    print("ESUNAccount：", Esun.main_account)

    personal_balance_sheet = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, "//a[text()='個人資產負債表']"))
    )
    driver.execute_script("arguments[0].click();", personal_balance_sheet)

    balance_td = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "fms01010a:twTd2"))
    )

    balance_text = balance_td.text.strip().replace(",", "")
    Esun.cash = int(balance_text)
    print("ESUNcash:", Esun.cash)

    balance_td = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "fms01010a:stockTd2"))
    )

    balance_text = balance_td.text.strip().replace(",", "")
    Esun.stock = int(balance_text)
    print("ESUNstock:", Esun.stock)

    logout_button = driver.find_element(By.CSS_SELECTOR, "a.log_out")  # 使用CSS選擇器定位
    logout_button.click()
def CathaySpider():
    driver.get("https://www.cathaybk.com.tw/mybank/")
    wait = WebDriverWait(driver, 120)

    cust_input = WebDriverWait(driver, 20).until(
        EC.visibility_of_element_located((By.ID, "CustID"))
    )
    driver.execute_script("arguments[0].value = arguments[1];", cust_input, Cathay.login_id)

    cust_input = WebDriverWait(driver, 20).until(
        EC.visibility_of_element_located((By.ID, "UserIdKeyin"))
    )
    driver.execute_script("arguments[0].value = arguments[1];", cust_input, Cathay.login_account)

    cust_input = WebDriverWait(driver, 20).until(
        EC.visibility_of_element_located((By.ID, "PasswordKeyin"))
    )
    driver.execute_script("arguments[0].value = arguments[1];", cust_input, Cathay.login_password)

    loginButton = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.XPATH, "//button[@type='button' and @class='btn no-print btn-fill js-login btn btn-fill w-100 u-pos-relative' and @onclick='NormalDataCheck()']"))
    )
    driver.execute_script("arguments[0].click();", loginButton)
    time.sleep(10)

    ###TWD
    button_element = WebDriverWait(driver, 20).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, 'button[data-evt="home_twd_overview"]'))
    )
    raw_text = button_element.text
    clean_text = raw_text.replace("TWD", "").replace(",", "").strip()
    Cathay.cash = int(clean_text)
    print(f"CATHAY_TWD: {Cathay.cash}")

    ###Foreign
    foreign_currency_element = WebDriverWait(driver, 20).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, 'button[data-evt="home_foreign_currency_overview"]'))
    )
    foreign_currency_text = foreign_currency_element.text
    clean_text = foreign_currency_text.replace("TWD", "").replace(",", "").strip().split()[0]
    Cathay.exchange = int(clean_text)
    print("CATHAYForeign:", Cathay.exchange)

    ###STOCK
    xpath_selector = "//p[text()='投資']/parent::div/following-sibling::div[@class='css-iu1euh']/p"
    
    investment_element = WebDriverWait(driver, 20).until(
        EC.visibility_of_element_located((By.XPATH, xpath_selector))
    )
    
    # 2. 獲取文字並清理
    investment_text = investment_element.text
    clean_text = investment_text.replace("TWD", "").replace(",", "").strip().split()[0]
    Cathay.stock = int(clean_text)
    print("CATHAYstock:", Cathay.stock)

    ###LOGOUT
    logout_button = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[data-evt="onlinebanking-logout"]'))
    )
    
    logout_button.click()


def LineSpider():
    driver.get("https://accessibility.linebank.com.tw/transaction")
    wait = WebDriverWait(driver, 20)

    wait.until(EC.presence_of_element_located((By.ID, "nationalId"))).send_keys(Line.login_id)
    wait.until(EC.presence_of_element_located((By.ID, "userId"))).send_keys(Line.login_account)
    wait.until(EC.presence_of_element_located((By.ID, "pw"))).send_keys(Line.login_password)

    login_btn = wait.until(
        EC.element_to_be_clickable((By.XPATH, "//button[@title='登入友善網路銀行']"))
    )
    login_btn.click()

    confirm_btn = wait.until(
        EC.element_to_be_clickable((By.XPATH, "//button[@title='確定']"))
    )
    confirm_btn.click()

    wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(., '可用餘額')]")))

    h2 = driver.find_element(By.XPATH, "//h2[contains(., '主帳戶')]")
    txt = re.sub(r"\s+", "", h2.text)                    
    Line.main_account = re.search(r"[（(]([0-9\-]+)[)）]", txt).group(1)
    print("LINEAccount:", Line.main_account)

    p = driver.find_element(By.XPATH, "//p[contains(., '可用餘額')]")
    ptxt = re.sub(r"\s+", "", p.text)                    
    m = re.search(r"NT\$?([0-9,]+(?:\.[0-9]+)?)", ptxt)
    available_display = f"NT${m.group(1)}"
    Line.cash = int(m.group(1).replace(",", ""))
    print("LINEcash:", Line.cash)
def JudgeColor(SheetRow, row):
    if SheetRow < 0:
        sheet.format(row, {'backgroundColor': {'red': 1, 'green': 0, 'blue': 0}})
    elif SheetRow > 0:
        sheet.format(row, {'backgroundColor': {'red': 0, 'green': 1, 'blue': 0}})
    elif SheetRow == 0:
        sheet.format(row, {'backgroundColor': {'red': 1, 'green': 1, 'blue': 1}})


EsunSpider()
CathaySpider()
LineSpider()

total_cash = Esun.cash + Cathay.cash + Line.cash
total_exchange = Cathay.exchange
total_stock = Esun.stock + Cathay.stock
total_assets = total_cash + total_exchange + total_stock
print("total_cash:", total_cash)
print("total_exchange:", total_exchange)
print("total_stock:", total_stock)
print("total_assets:", total_assets)

C3_value = int(sheet.cell(3, 3).value)
D3_value = int(sheet.cell(3, 4).value) 
E3_value = int(sheet.cell(3, 5).value) 
F3_value = int(sheet.cell(3, 6).value) 

cash_diff = total_cash - C3_value
exchange_diff = total_exchange - D3_value
stock_diff = total_stock - E3_value
assets_diff = total_assets - F3_value

current_date = datetime.now().strftime("%Y/%m/%d")
current_time = datetime.now().strftime("%H:%M:%S")
current_date2 = datetime.now().strftime("%m/%d")

sheet.insert_row([current_date, current_time, 
                total_cash, total_exchange, total_stock, total_assets, 
                cash_diff, exchange_diff, stock_diff, assets_diff, " ", 
                Esun.main_account, Esun.cash, Esun.exchange, Esun.stock, " ",  
                Cathay.main_account, Cathay.cash, Cathay.exchange, Cathay.stock, " ",  
                Line.main_account, Line.cash, Line.exchange, Line.stock
                ], 3)

G3_value = int(sheet.cell(3, 7).value)   
H3_value = int(sheet.cell(3, 8).value)
I3_value = int(sheet.cell(3, 9).value)
J3_value = int(sheet.cell(3, 10).value)

JudgeColor(G3_value, 'G3')
JudgeColor(H3_value, 'H3')
JudgeColor(I3_value, 'I3')
JudgeColor(J3_value, 'J3')
